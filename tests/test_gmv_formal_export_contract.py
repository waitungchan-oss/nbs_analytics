from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from app_workflows import (
    _apply_gmv_refund_adjustments,
    _build_gmv_audit_workbook,
    _gmv_summary_rows,
    build_formal_gmv_workbooks,
)
from tests.fixtures.gmv_export_semantic_fixture import read_gmv_workbook_semantics, semantic_fixture


def test_formal_export_keeps_total_paid_and_official_net_products():
    tour = pd.DataFrame(
        [{"來源單據號": "S-1", "收款原幣金額": 100.0, "數量": 2}]
    )
    refunds = pd.DataFrame(
        [
            {"退款單號": "R-1", "來源單據號": "S-1", "退款原幣金額": 20.0, "退款狀態": "退款中"},
            {"退款單號": "R-2", "來源單據號": "S-1", "退款原幣金額": 30.0, "退款狀態": "已退款"},
        ]
    )
    total = _apply_gmv_refund_adjustments(tour, pd.DataFrame(), refunds)
    paid = _apply_gmv_refund_adjustments(tour, pd.DataFrame(), refunds, refund_status="已退款")
    workbooks = build_formal_gmv_workbooks(
        total_adjusted=total,
        paid_adjusted=paid,
        total_summary_rows=_gmv_summary_rows(tour, pd.DataFrame(), total),
        paid_summary_rows=_gmv_summary_rows(tour, pd.DataFrame(), paid),
        provenance={"version_id": "V-1", "quantity_basis": "原交易人數／數量（未按退款調整）"},
    )

    assert set(workbooks) == {"total", "paid", "formal"}
    total_sheets = load_workbook(BytesIO(workbooks["total"]), read_only=True).sheetnames
    paid_sheets = load_workbook(BytesIO(workbooks["paid"]), read_only=True).sheetnames
    formal_sheets = load_workbook(BytesIO(workbooks["formal"]), read_only=True).sheetnames
    assert "總退款摘要" in total_sheets
    assert "已退款摘要" in paid_sheets
    assert "正式淨GMV摘要" in formal_sheets


def test_formal_audit_workbook_contains_provenance_and_original_quantity_label():
    detail = pd.DataFrame(
        [{"來源單據號": "S-1", "退款扣減金額": 30.0, "數量": 2}]
    )
    payload = _build_gmv_audit_workbook(
        [{"退款維度": "已退款", "指標": "實際扣減金額", "數值": 30.0}],
        detail,
        [],
        dimension="正式淨GMV",
        provenance={"version_id": "V-1", "quantity_basis": "原交易人數／數量（未按退款調整）"},
    )
    sheets = load_workbook(BytesIO(payload), read_only=True)

    assert "正式淨GMV摘要" in sheets.sheetnames
    assert "Provenance" in sheets.sheetnames
    values = [cell.value for row in sheets["Provenance"].iter_rows() for cell in row]
    assert "V-1" in values
    assert "原交易人數／數量（未按退款調整）" in values


def test_semantic_reader_covers_legacy_audit_workbook_contract():
    _, _, formal_tour, formal_others, refunds = semantic_fixture()
    total = _apply_gmv_refund_adjustments(formal_tour, formal_others, refunds)
    paid = _apply_gmv_refund_adjustments(formal_tour, formal_others, refunds, refund_status="已退款")
    workbooks = build_formal_gmv_workbooks(
        total_adjusted=total,
        paid_adjusted=paid,
        total_summary_rows=_gmv_summary_rows(formal_tour, formal_others, total),
        paid_summary_rows=_gmv_summary_rows(formal_tour, formal_others, paid),
        provenance={"version_id": "fixture-v1"},
    )
    semantics = read_gmv_workbook_semantics(workbooks["paid"])
    assert semantics["sheetNames"] == ["已退款摘要", "已退款扣減明細", "已退款未匹配來源單據號", "Provenance"]
    assert semantics["sheets"]["已退款摘要"]["headers"] == ["退款維度", "指標", "數值"]
    assert semantics["sheets"]["已退款摘要"]["stableKeys"][0] == {"退款維度": "已退款", "指標": "退款來源訂單數"}
    assert semantics["sheets"]["已退款扣減明細"]["rowCount"] == 3
    assert any("已退款" in row for row in semantics["sheets"]["Provenance"]["rows"])
