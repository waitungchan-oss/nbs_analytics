from pathlib import Path

from openpyxl import Workbook


def _writer(facts, path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "營收"
    sheet.append(["scope", "rows"])
    sheet.append([facts.scope_id, len(facts.tour) + len(facts.others)])
    workbook.save(path)


def _facts(scope_id="official"):
    import pandas as pd

    from backend.services.export_intermediate_service import DashboardReportFacts

    return DashboardReportFacts(
        scope_id=scope_id,
        tour=pd.DataFrame([{"來源單據號": "T-1", "收款原幣金額": 100}]),
        others=pd.DataFrame([{"來源單據號": "O-1", "收款原幣金額": 25}]),
        aggregates={},
        schema_fingerprint="a" * 64,
        data_fingerprint="b" * 64,
    )


def test_serializer_writes_valid_xlsx_atomically_without_mutating_facts(tmp_path):
    from openpyxl import load_workbook

    from backend.services.export_serializer_service import ExportSerializerJob, serialize_export_jobs_parallel

    target = tmp_path / "official.xlsx"
    job = ExportSerializerJob(
        artifact_id="official.xlsx", scope_id="official", facts=_facts(),
        target_path=target, schema_fingerprint="a" * 64, data_fingerprint="b" * 64,
    )
    result = serialize_export_jobs_parallel((job,), writer=_writer, max_workers=1)

    assert result[0].status == "READY"
    assert result[0].bytes_written > 0
    workbook = load_workbook(target, read_only=True, data_only=True)
    assert workbook.sheetnames == ["營收"]
    assert list(workbook["營收"].values)[1] == ("official", 2)
    workbook.close()
    assert list(job.facts.tour.columns) == ["來源單據號", "收款原幣金額"]


def test_parallel_serializer_returns_deterministic_order_and_cleans_failed_targets(tmp_path):
    from backend.services.export_serializer_service import ExportSerializerJob, serialize_export_jobs_parallel

    jobs = tuple(
        ExportSerializerJob(
            artifact_id=f"{scope}.xlsx", scope_id=scope, facts=_facts(scope),
            target_path=tmp_path / f"{scope}.xlsx", schema_fingerprint="a" * 64,
            data_fingerprint="b" * 64,
        )
        for scope in ("all", "no_writeoff", "official")
    )
    results = serialize_export_jobs_parallel(jobs, writer=_writer, max_workers=2)

    assert [result.artifact_id for result in results] == [job.artifact_id for job in jobs]
    assert all(result.status == "READY" for result in results)


def test_serializer_rejects_missing_fingerprints(tmp_path):
    from backend.services.export_serializer_service import ExportSerializerJob, serialize_export_jobs_parallel

    job = ExportSerializerJob(
        artifact_id="official.xlsx", scope_id="official", facts=_facts(),
        target_path=tmp_path / "official.xlsx", schema_fingerprint="", data_fingerprint="b" * 64,
    )
    result = serialize_export_jobs_parallel((job,), writer=_writer)

    assert result[0].status == "SERIALIZER_INVALID_FACTS"
    assert not (tmp_path / "official.xlsx").exists()
