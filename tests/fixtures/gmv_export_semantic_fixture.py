"""Small, explicit GMV export fixture covering the dual refund dimensions."""

from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import load_workbook


def semantic_fixture() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Return raw/formal tour and ticket frames plus refund observations.

    The raw rows deliberately include scope-excluded revenue while the formal
    frames model the production revenue scope after excluding 掛賬核銷 and
    TT 退款轉團款.
    """
    tour = pd.DataFrame(
        [
            {"來源單據號": "T-ORD", "統一日期": "2026-05-01", "銷售點": "銅鑼灣分社", "銷售員": "Alice", "收款原幣金額": 100, "收款類型": "正常收款", "收款方式": "現金", "數量": 2, "團負責人部門": ""},
            {"來源單據號": "T-EXCLUDED", "統一日期": "2026-05-01", "銷售點": "銅鑼灣分社", "銷售員": "Bob", "收款原幣金額": 80, "收款類型": "掛賬核銷", "收款方式": "現金", "數量": 1, "團負責人部門": ""},
            {"來源單據號": "T-TT", "統一日期": "2026-05-01", "銷售點": "銅鑼灣分社", "銷售員": "Carol", "收款原幣金額": 90, "收款類型": "正常收款", "收款方式": "TT 退款轉團款", "數量": 1, "團負責人部門": ""},
            {"來源單據號": "T-PART", "統一日期": "2026-05-02", "銷售點": "上環服務點", "銷售員": "Alice", "收款原幣金額": 120, "收款類型": "正常收款", "收款方式": "信用卡", "數量": 3, "團負責人部門": ""},
            {"來源單據號": "T-OVER", "統一日期": "2026-05-02", "銷售點": "上環服務點", "銷售員": "Ben", "收款原幣金額": 50, "收款類型": "正常收款", "收款方式": "信用卡", "數量": 1, "團負責人部門": "郵輪部"},
            {"來源單據號": "T-ZERO", "統一日期": "2026-05-02", "銷售點": "元朗服務點", "銷售員": "Ben", "收款原幣金額": 0, "收款類型": "正常收款", "收款方式": "", "數量": 0, "團負責人部門": ""},
            {"來源單據號": "T-DUP", "統一日期": "2026-05-03", "銷售點": "元朗服務點", "銷售員": "Specialist", "收款原幣金額": 30, "收款類型": "正常收款", "收款方式": "", "數量": 1, "團負責人部門": "專職銷售組"},
        ]
    )
    others = pd.DataFrame(
        [
            {"來源單據號": "O-TICKET", "統一日期": "2026-05-03", "銷售點": "銅鑼灣分社", "銷售員": "Alice", "收款原幣金額": 40, "收款類型": "正常收款", "收款方式": "信用卡", "數量": 4, "來源報表標籤": "門券all", "團名稱": "景點門票"},
        ]
    )
    formal_tour = tour.loc[
        (tour["收款類型"] != "掛賬核銷") & (tour["收款方式"] != "TT 退款轉團款")
    ].copy()
    formal_others = others.copy()
    refunds = pd.DataFrame(
        [
            {"退款單號": "R-ORD", "來源單據號": "T-ORD", "退款原幣金額": 10, "退款狀態": "待退款", "退款方式": "現金"},
            {"退款單號": "R-TT", "來源單據號": "T-TT", "退款原幣金額": 90, "退款狀態": "待退款", "退款方式": "TT 退款轉團款"},
            {"退款單號": "R-PART", "來源單據號": "T-PART", "退款原幣金額": 20, "退款狀態": "已退款", "退款方式": "信用卡"},
            {"退款單號": "R-OVER", "來源單據號": "T-OVER", "退款原幣金額": 80, "退款狀態": "已退款", "退款方式": "信用卡"},
            {"退款單號": "R-ZERO", "來源單據號": "T-ZERO", "退款原幣金額": 0, "退款狀態": "已退款", "退款方式": "信用卡"},
            {"退款單號": "R-DUP-1", "來源單據號": "T-DUP", "退款原幣金額": 10, "退款狀態": "已退款", "退款方式": "信用卡"},
            {"退款單號": "R-DUP-2", "來源單據號": "T-DUP", "退款原幣金額": 5, "退款狀態": "已退款", "退款方式": "信用卡"},
            {"退款單號": "R-MISSING", "來源單據號": "MISSING", "退款原幣金額": 7, "退款狀態": "已退款", "退款方式": "信用卡"},
        ]
    )
    return tour, others, formal_tour, formal_others, refunds


def _normalized(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return format(value, ".12g")
    return str(value).strip()


def read_gmv_workbook_semantics(content: bytes) -> dict[str, object]:
    """Extract stable workbook semantics without relying on XLSX bytes."""
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheets: dict[str, dict[str, object]] = {}
    for sheet in workbook.worksheets:
        rows = [[_normalized(value) for value in row] for row in sheet.iter_rows(values_only=True)]
        headers = rows[0] if rows else []
        key_indexes = {
            header: index
            for index, header in enumerate(headers)
            if header in {"來源單據號", "退款維度", "指標", "欄位"}
        }
        stable_keys = [
            {header: row[index] for header, index in key_indexes.items() if index < len(row)}
            for row in rows[1:]
        ]
        sheets[sheet.title] = {
            "headers": headers,
            "rowCount": max(len(rows) - 1, 0),
            "rows": rows[1:],
            "stableKeys": stable_keys,
        }
    workbook.close()
    return {"sheetNames": list(sheets), "sheets": sheets}
