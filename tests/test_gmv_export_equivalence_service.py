import json
from io import BytesIO

import pandas as pd
import pytest
from openpyxl import Workbook


def _xlsx(amount=100, sheet="營收"):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet
    worksheet.append(["來源單據號", "收款原幣金額", "數量"])
    worksheet.append(["A-1", amount, 2])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _provenance_xlsx(*, version_id: str, created_at: str, amount: float = 100) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "營收"
    worksheet.append(["來源單據號", "收款原幣金額"])
    worksheet.append(["A-1", amount])
    provenance = workbook.create_sheet("Provenance")
    provenance.append(["欄位", "值"])
    provenance.append(["version_id", version_id])
    provenance.append(["generated_at", created_at])
    provenance.append(["quantity_basis", "原交易人數／數量（未按退款調整）"])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _semantic_artifacts() -> tuple[dict[str, bytes], dict[str, str]]:
    from backend.services.gmv_export_cache_service import CANONICAL_CACHE_ARTIFACT_KEYS

    artifacts = {}
    kinds = {}
    for key in CANONICAL_CACHE_ARTIFACT_KEYS:
        if key.endswith(".detail"):
            artifacts[key] = pd.DataFrame([{"來源單據號": "A-1", "收款原幣金額": 100.0}]).to_csv(index=False).encode()
            kinds[key] = "csv"
        elif key == "summaries":
            artifacts[key] = json.dumps([{"退款維度": "總退款", "指標": "實際扣減金額", "數值": 100.0}], ensure_ascii=False).encode()
            kinds[key] = "json"
        else:
            artifacts[key] = _xlsx()
            kinds[key] = "xlsx"
    return artifacts, kinds


def test_gmv_workbook_equivalence_normalizes_money_and_metadata():
    from backend.services.gmv_export_equivalence_service import compare_gmv_workbook_semantics

    result = compare_gmv_workbook_semantics(
        reference_bytes=_xlsx(100.004), candidate_bytes=_xlsx(100.00), artifact_id="total.ex.xlsx",
    )
    assert result.status == "PASS"
    assert result.mismatch_count == 0


def test_gmv_workbook_equivalence_reports_schema_and_value_mismatch():
    from backend.services.gmv_export_equivalence_service import compare_gmv_workbook_semantics

    result = compare_gmv_workbook_semantics(
        reference_bytes=_xlsx(100), candidate_bytes=_xlsx(101, sheet="不同"), artifact_id="paid.ex.xlsx",
    )
    assert result.status == "FAIL"
    assert result.mismatch_count >= 1
    assert len(result.mismatch_examples) <= 20


def test_gmv_artifact_equivalence_requires_exact_keys_and_valid_dimensions():
    from backend.services.gmv_export_equivalence_service import CANONICAL_ARTIFACT_IDS, compare_gmv_artifact_sets

    reference = {key: _xlsx() for key in CANONICAL_ARTIFACT_IDS}
    candidate = dict(reference)
    missing = compare_gmv_artifact_sets(
        reference=reference, candidate={key: value for key, value in candidate.items() if key != "paid.detail"}
    )
    assert missing.status == "FAIL"
    assert missing.mismatch_count == 1

    try:
        compare_gmv_artifact_sets(reference={"unknown.official": _xlsx()}, candidate={"unknown.official": _xlsx()})
    except ValueError as exc:
        assert "artifact" in str(exc)
    else:
        raise AssertionError("unknown GMV artifact dimension must fail closed")


def test_semantic_records_cover_exact_artifact_contract_and_are_compact():
    from backend.services.gmv_export_cache_service import CANONICAL_CACHE_ARTIFACT_KEYS
    from backend.services.gmv_export_equivalence_service import (
        build_gmv_artifact_semantic_records,
        compare_gmv_artifact_semantics,
    )

    artifacts, kinds = _semantic_artifacts()
    records = build_gmv_artifact_semantic_records(artifacts, kinds)

    assert tuple(records) == tuple(sorted(CANONICAL_CACHE_ARTIFACT_KEYS))
    assert all(set(record) == {
        "kind", "schemaFingerprint", "semanticFingerprint", "rowCount", "sheetCount",
    } for record in records.values())
    assert records["summaries"]["kind"] == "json"
    assert records["total.detail"]["kind"] == "csv"
    assert records["total.workbook.ex.xlsx"]["kind"] == "xlsx"
    assert compare_gmv_artifact_semantics(records, records).status == "PASS"


def test_semantic_records_ignore_dynamic_provenance_but_detect_money_change():
    from backend.services.gmv_export_equivalence_service import build_gmv_artifact_semantic_records

    reference = {"total.workbook.audit.xlsx": _provenance_xlsx(version_id="v1", created_at="2026-08-25T01:00:00Z")}
    candidate = {"total.workbook.audit.xlsx": _provenance_xlsx(version_id="v2", created_at="2026-08-25T02:00:00Z")}
    kinds = {"total.workbook.audit.xlsx": "xlsx"}
    reference_record = build_gmv_artifact_semantic_records(reference, kinds)
    candidate_record = build_gmv_artifact_semantic_records(candidate, kinds)
    assert reference_record["total.workbook.audit.xlsx"]["semanticFingerprint"] == candidate_record["total.workbook.audit.xlsx"]["semanticFingerprint"]

    changed = {"total.workbook.audit.xlsx": _provenance_xlsx(version_id="v2", created_at="2026-08-25T02:00:00Z", amount=101)}
    changed_record = build_gmv_artifact_semantic_records(changed, kinds)
    assert changed_record["total.workbook.audit.xlsx"]["semanticFingerprint"] != reference_record["total.workbook.audit.xlsx"]["semanticFingerprint"]


def test_semantic_comparison_detects_schema_row_and_artifact_key_changes():
    from backend.services.gmv_export_equivalence_service import (
        build_gmv_artifact_semantic_records,
        compare_gmv_artifact_semantics,
    )

    reference_bytes = {"total.workbook.ex.xlsx": _xlsx()}
    candidate_bytes = {"total.workbook.ex.xlsx": _xlsx(amount=101)}
    kinds = {"total.workbook.ex.xlsx": "xlsx"}
    reference = build_gmv_artifact_semantic_records(reference_bytes, kinds)
    candidate = build_gmv_artifact_semantic_records(candidate_bytes, kinds)
    mismatch = compare_gmv_artifact_semantics(reference, candidate)
    assert mismatch.status == "FAIL"
    assert mismatch.mismatch_count >= 1

    missing = compare_gmv_artifact_semantics(reference, {})
    assert missing.status == "FAIL"
    assert any(example.get("kind") == "artifact_keys" for example in missing.mismatch_examples)


def test_source_identity_is_separate_from_semantic_shadow_identity():
    from backend.services.gmv_export_equivalence_service import (
        build_gmv_artifact_semantic_records,
        compare_gmv_artifact_semantics,
    )
    from backend.services.gmv_trusted_reference_service import build_gmv_content_fingerprint

    source_kwargs = {
        "revenue_generation_token": "revenue-v1",
        "refund_state_sha256": "a" * 64,
        "rule_version": "不含掛賬核銷與TT退款轉團款",
        "export_schema_version": "gmv-formal-export-v2",
        "pipeline_fingerprint": "pipeline-gmv-fast-v1",
        "serializer_version": "gmv-openpyxl-serializer-v1",
    }
    source_identity = build_gmv_content_fingerprint(**source_kwargs)
    reference_bytes, kinds = _semantic_artifacts()
    candidate_bytes = dict(reference_bytes)
    candidate_bytes["total.detail"] = pd.DataFrame(
        [{"來源單據號": "A-1", "收款原幣金額": 101.0}],
    ).to_csv(index=False).encode()

    reference = build_gmv_artifact_semantic_records(reference_bytes, kinds)
    candidate = build_gmv_artifact_semantic_records(candidate_bytes, kinds)

    assert build_gmv_content_fingerprint(**source_kwargs) == source_identity
    comparison = compare_gmv_artifact_semantics(reference, candidate)
    assert comparison.status == "FAIL"
    assert any(example.get("artifact") == "total.detail" for example in comparison.mismatch_examples)


def test_semantic_record_builder_rejects_wrong_kind_and_unknown_keys():
    from backend.services.gmv_export_equivalence_service import build_gmv_artifact_semantic_records

    with pytest.raises(ValueError, match="artifact"):
        build_gmv_artifact_semantic_records(
            {"unknown": b"x"}, {"unknown": "json"},
        )
    with pytest.raises(ValueError, match="kind"):
        build_gmv_artifact_semantic_records(
            {"total.detail": b"x"}, {"total.detail": "xlsx"},
        )


def test_json_semantics_ignore_provenance_and_normalize_numeric_summary_values():
    from backend.services.gmv_export_equivalence_service import build_gmv_artifact_semantic_records

    reference = {
        "summaries": json.dumps({
            "rows": [{"退款維度": "總退款", "數值": 100.0}],
            "provenance": {"version_id": "v1", "generated_at": "2026-08-25T01:00:00Z"},
        }, ensure_ascii=False).encode()
    }
    candidate = {
        "summaries": json.dumps({
            "rows": [{"退款維度": "總退款", "數值": "100.00"}],
            "provenance": {"version_id": "v2", "created_at": "2026-08-25T02:00:00Z"},
        }, ensure_ascii=False).encode()
    }
    kinds = {"summaries": "json"}
    reference_record = build_gmv_artifact_semantic_records(reference, kinds)
    candidate_record = build_gmv_artifact_semantic_records(candidate, kinds)
    assert reference_record["summaries"]["semanticFingerprint"] == candidate_record["summaries"]["semanticFingerprint"]

    changed = {
        "summaries": json.dumps({
            "rows": [{"退款維度": "總退款", "數值": 101}],
            "provenance": {"version_id": "v3"},
        }, ensure_ascii=False).encode()
    }
    changed_record = build_gmv_artifact_semantic_records(changed, kinds)
    assert changed_record["summaries"]["semanticFingerprint"] != reference_record["summaries"]["semanticFingerprint"]


def test_json_semantics_keep_dynamic_labels_in_business_rows():
    from backend.services.gmv_export_equivalence_service import build_gmv_artifact_semantic_records

    kinds = {"summaries": "json"}
    reference = {"summaries": json.dumps([{"退款維度": "總退款", "created_at": "2026-08-25"}], ensure_ascii=False).encode()}
    candidate = {"summaries": json.dumps([{"退款維度": "總退款", "created_at": "2026-08-26"}], ensure_ascii=False).encode()}

    reference_record = build_gmv_artifact_semantic_records(reference, kinds)
    candidate_record = build_gmv_artifact_semantic_records(candidate, kinds)
    assert reference_record["summaries"]["semanticFingerprint"] != candidate_record["summaries"]["semanticFingerprint"]


def test_json_semantics_preserve_business_visible_array_order():
    from backend.services.gmv_export_equivalence_service import build_gmv_artifact_semantic_records

    kinds = {"summaries": "json"}
    reference = {"summaries": json.dumps([
        {"退款維度": "總退款", "數值": 100},
        {"退款維度": "已退款", "數值": 80},
    ], ensure_ascii=False).encode()}
    candidate = {"summaries": json.dumps([
        {"退款維度": "已退款", "數值": 80},
        {"退款維度": "總退款", "數值": 100},
    ], ensure_ascii=False).encode()}

    reference_record = build_gmv_artifact_semantic_records(reference, kinds)
    candidate_record = build_gmv_artifact_semantic_records(candidate, kinds)
    assert reference_record["summaries"]["semanticFingerprint"] != candidate_record["summaries"]["semanticFingerprint"]
