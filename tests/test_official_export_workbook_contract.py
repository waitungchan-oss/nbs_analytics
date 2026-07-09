from io import BytesIO

import pandas as pd
from openpyxl import load_workbook


def _sheet_rows(workbook, sheet_name):
    return list(workbook[sheet_name].iter_rows(values_only=True))


def _sum_rows(rows, columns):
    return sum(sum(float(row[index] or 0) for index in columns) for row in rows[1:])


def test_official_export_branch_salesperson_sheet_matches_existing_branch_sheets():
    import pipeline

    tour = pd.DataFrame(
        [
            {
                "來源單據號": "T001",
                "統一日期": "2026-05-01",
                "銷售點": "銅鑼灣分社",
                "銷售員": "Alice",
                "收款原幣金額": 100,
                "收款類型": "正常收款",
                "收款方式": "現金",
                "團負責人部門": "",
                "交易時間": "2026-05-01",
                "行程天數": 3,
                "數量": 2,
            },
            {
                "來源單據號": "T002",
                "統一日期": "2026-05-01",
                "銷售點": "銅鑼灣分社",
                "銷售員": "Ben",
                "收款原幣金額": 50,
                "收款類型": "正常收款",
                "收款方式": "現金",
                "團負責人部門": "郵輪部",
                "交易時間": "2026-05-01",
                "行程天數": 5,
                "數量": 1,
            },
            {
                "來源單據號": "T999",
                "統一日期": "2026-05-01",
                "銷售點": "銅鑼灣分社",
                "銷售員": "Alice",
                "收款原幣金額": 999,
                "收款類型": "掛賬核銷",
                "收款方式": "現金",
                "團負責人部門": "",
                "交易時間": "2026-05-01",
                "行程天數": 3,
                "數量": 9,
            },
        ]
    )
    others = pd.DataFrame(
        [
            {
                "來源單據號": "O001",
                "統一日期": "2026-05-01",
                "銷售點": "銅鑼灣分社",
                "銷售員": "Alice",
                "收款原幣金額": 25,
                "收款類型": "正常收款",
                "收款方式": "信用卡",
                "交易時間": "2026-05-01",
                "團名稱": "景點門票",
                "來源報表標籤": "門券all",
                "行程天數": 0,
                "數量": 3,
            },
            {
                "來源單據號": "O002",
                "統一日期": "2026-05-01",
                "銷售點": "銅鑼灣分社",
                "銷售員": "Ben",
                "收款原幣金額": 75,
                "收款類型": "正常收款",
                "收款方式": "信用卡",
                "交易時間": "2026-05-01",
                "團名稱": "景點門票",
                "來源報表標籤": "門券all",
                "行程天數": 0,
                "數量": 4,
            },
            {
                "來源單據號": "O999",
                "統一日期": "2026-05-01",
                "銷售點": "銅鑼灣分社",
                "銷售員": "Ben",
                "收款原幣金額": 888,
                "收款類型": "正常收款",
                "收款方式": "TT 退款轉團款",
                "交易時間": "2026-05-01",
                "團名稱": "景點門票",
                "來源報表標籤": "門券all",
                "行程天數": 0,
                "數量": 8,
            },
        ]
    )

    workbook_buffer, _, _ = pipeline.build_dashboard_data_excluding_receipt_types(
        tour,
        others,
        {"1A": "銅鑼灣分社", "225": "專職銷售組"},
        ["銅鑼灣分社"],
        ["郵輪部"],
        ["Alice", "Ben"],
        ["掛賬核銷"],
        excluded_payment_methods=["TT 退款轉團款"],
        include_branch_salesperson_sheet=True,
    )

    workbook = load_workbook(BytesIO(workbook_buffer.getvalue()), data_only=True)
    for sheet_name in [
        "分社經營統計",
        "分社經營統計_含銷售員",
        "分社每天旅行團交易人數",
        "分社每天票務交易數量",
    ]:
        assert sheet_name in workbook.sheetnames

    branch_rows = _sheet_rows(workbook, "分社經營統計")
    salesperson_rows = _sheet_rows(workbook, "分社經營統計_含銷售員")
    tour_people_rows = _sheet_rows(workbook, "分社每天旅行團交易人數")
    ticket_quantity_rows = _sheet_rows(workbook, "分社每天票務交易數量")

    assert salesperson_rows[0] == (
        "文本",
        "單選",
        "銷售員",
        "日期",
        "月份",
        "旅行團",
        "郵輪",
        "票務",
        "旅行團交易人數",
        "票務交易數量",
    )
    assert _sum_rows(branch_rows, [4, 5, 6]) == 250.0
    assert _sum_rows(salesperson_rows, [5, 6, 7]) == 250.0
    assert _sum_rows(tour_people_rows, [3, 4]) == 3.0
    assert _sum_rows(salesperson_rows, [8]) == 3.0
    assert _sum_rows(ticket_quantity_rows, [4]) == 7.0
    assert _sum_rows(salesperson_rows, [9]) == 7.0
