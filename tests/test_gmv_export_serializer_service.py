from pathlib import Path
import time

import pandas as pd


def _facts(dimension="總退款", scope_id="official"):
    from backend.services.gmv_export_intermediate_service import GmvReportFacts

    return GmvReportFacts(
        dimension=dimension,
        scope_id=scope_id,
        sheets={
            "營收": pd.DataFrame([{"來源單據號": "A-1", "收款原幣金額": 100.0}]),
            "摘要": pd.DataFrame([{"指標": "GMV", "數值": 100.0}]),
        },
        row_counts={"營收": 1, "摘要": 1},
        schema_fingerprint="schema",
        data_fingerprint="data",
    )


def _gate():
    from backend.services.gmv_export_serializer_service import SerializerPublicationGate
    return SerializerPublicationGate("PASS", "PASS", "PASS", "PASS", "PASS")


def test_serializer_publication_gate_rejects_shadow_mismatch():
    from backend.services.gmv_export_serializer_service import SerializerPublicationGate

    assert not SerializerPublicationGate("PASS", "PASS", "PASS", "PASS", "MISMATCH").ready
    assert not SerializerPublicationGate("PASS", "PASS", "PASS", "FAIL", "PASS").ready


def test_serializer_staging_gate_allows_pending_shadow_only_for_private_staging():
    from backend.services.gmv_export_serializer_service import SerializerPublicationGate

    assert SerializerPublicationGate("PENDING", "PASS", "PASS", "PASS", "PENDING", staging_only=True).ready
    assert not SerializerPublicationGate("PASS", "PASS", "PASS", "PASS", "PENDING", staging_only=True).ready
    assert not SerializerPublicationGate("PASS", "PASS", "PASS", "PASS", "PENDING").ready


def test_bounded_serializer_timeout_scales_large_batches_but_has_a_cap():
    from backend.services.gmv_export_serializer_service import (
        SerializerJob, bounded_serializer_timeout_seconds,
    )

    small = [SerializerJob("total.official", _facts(), Path("total.xlsx"), _gate())]
    large = [
        SerializerJob(
            f"total.{index}",
            type(_facts())(
                dimension="總退款", scope_id="official", sheets={},
                row_counts={"rows": 500_000}, schema_fingerprint="schema", data_fingerprint="data",
            ),
            Path(f"total-{index}.xlsx"), _gate(),
        )
        for index in range(6)
    ]

    assert bounded_serializer_timeout_seconds(small) == 120
    assert bounded_serializer_timeout_seconds(large) == 300


def test_serializer_writes_valid_xlsx_atomically_without_mutating_facts(tmp_path):
    from openpyxl import load_workbook
    from backend.services.gmv_export_serializer_service import serialize_gmv_report_facts

    facts = _facts()
    before = {name: frame.copy(deep=True) for name, frame in facts.sheets.items()}
    result = serialize_gmv_report_facts(
        facts, artifact_path=tmp_path / "official.xlsx", artifact_id="total.official", publication_gate=_gate(),
    )

    assert result.status == "READY"
    assert result.error is None
    assert result.path == tmp_path / "official.xlsx"
    workbook = load_workbook(result.path, read_only=True, data_only=True)
    assert workbook.sheetnames == ["營收", "摘要"]
    workbook.close()
    for name in before:
        pd.testing.assert_frame_equal(facts.sheets[name], before[name])


def test_parallel_serializer_is_bounded_and_returns_deterministic_order(tmp_path):
    from backend.services.gmv_export_serializer_service import SerializerJob, serialize_gmv_workbooks_parallel

    jobs = [
        SerializerJob("total.official", _facts(), tmp_path / "total.xlsx", _gate()),
        SerializerJob("paid.official", _facts("已退款"), tmp_path / "paid.xlsx", _gate()),
    ]
    results = serialize_gmv_workbooks_parallel(jobs, max_workers=2)

    assert [result.artifact_id for result in results] == ["total.official", "paid.official"]
    assert all(result.status == "READY" for result in results)


def test_serializer_reports_invalid_artifact_and_never_marks_ready(tmp_path):
    from backend.services.gmv_export_serializer_service import serialize_gmv_report_facts

    facts = _facts()
    result = serialize_gmv_report_facts(
        facts, artifact_path=tmp_path / "not-xlsx.txt", artifact_id="total.official", publication_gate=_gate(),
    )

    assert result.status == "SERIALIZER_INVALID_ARTIFACT"
    assert result.error


def test_serializer_rejects_unfingerprinted_facts_before_publication(tmp_path):
    from backend.services.gmv_export_intermediate_service import GmvReportFacts
    from backend.services.gmv_export_serializer_service import serialize_gmv_report_facts

    facts = GmvReportFacts(
        dimension="總退款",
        scope_id="official",
        sheets={"營收": pd.DataFrame([{"來源單據號": "A-1"}])},
        row_counts={"營收": 1},
        schema_fingerprint="",
        data_fingerprint="",
    )
    result = serialize_gmv_report_facts(
        facts, artifact_path=tmp_path / "invalid.xlsx", artifact_id="total.official", publication_gate=_gate(),
    )

    assert result.status == "SERIALIZER_INVALID_FACTS"
    assert not (tmp_path / "invalid.xlsx").exists()


def test_parallel_serializer_rejects_invalid_worker_count(tmp_path):
    from backend.services.gmv_export_serializer_service import SerializerJob, serialize_gmv_workbooks_parallel

    try:
        serialize_gmv_workbooks_parallel(
            [SerializerJob("total.official", _facts(), tmp_path / "total.xlsx", _gate())],
            max_workers=0,
        )
    except ValueError as exc:
        assert "max_workers" in str(exc)
    else:
        raise AssertionError("invalid max_workers must fail closed")


def test_parallel_serializer_reports_timeout(monkeypatch, tmp_path):
    import backend.services.gmv_export_serializer_service as service
    from backend.services.gmv_export_serializer_service import SerializerJob

    original = service.serialize_gmv_report_facts

    def slow(*args, **kwargs):
        time.sleep(0.05)
        return original(*args, **kwargs)

    monkeypatch.setattr(service, "serialize_gmv_report_facts", slow)
    result = service.serialize_gmv_workbooks_parallel(
        [SerializerJob("total.official", _facts(), tmp_path / "total.xlsx", _gate())],
        max_workers=1,
        timeout_seconds=0.001,
    )
    assert result[0].status == "SERIALIZER_TIMEOUT"
    time.sleep(0.08)
    assert not (tmp_path / "total.xlsx").exists()
