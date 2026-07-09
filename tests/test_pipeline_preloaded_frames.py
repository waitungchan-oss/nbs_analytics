import pandas as pd
from openpyxl import load_workbook


def test_read_excel_source_accepts_preloaded_dataframe_tuple_without_re_reading(monkeypatch):
    import pipeline

    def fail_read_excel(*args, **kwargs):
        raise AssertionError("pd.read_excel should not be called for preloaded frames")

    monkeypatch.setattr(pipeline.pd, "read_excel", fail_read_excel)
    source = pd.DataFrame({"來源單據號": ["17001"], "收款時間": ["2026-06-15"]})

    frame, name = pipeline._read_excel_source(("旅行團0628all.xlsx", source))

    assert name == "旅行團0628all.xlsx"
    assert frame.equals(source)
    assert frame is not source


def test_branch_reassignment_override_moves_2026_06_e6_to_0a_only():
    import pipeline

    source = pd.DataFrame(
        [
            {
                "來源單據號": "E6A2026001",
                "統一日期": "2026-06-15",
                "銷售點": "上環服務點",
                "副表_銷售點": "上環服務點",
            },
            {
                "來源單據號": "E6A2026071",
                "統一日期": "2026-07-15",
                "銷售點": "上環服務點",
                "副表_銷售點": "上環服務點",
            },
        ]
    )

    result = pipeline.apply_branch_reassignment_overrides(
        source,
        [
            {
                "month": "2026-06",
                "from_prefix": "E6",
                "from_branch": "上環服務點",
                "to_branch": "展覽會場專用",
                "to_prefix": "0A",
            }
        ],
    )

    assert result.loc[0, "銷售點"] == "展覽會場專用"
    assert result.loc[0, "副表_銷售點"] == "展覽會場專用"
    assert result.loc[1, "銷售點"] == "上環服務點"
    assert result.loc[1, "副表_銷售點"] == "上環服務點"


def test_export_workbook_adds_branch_summary_with_salesperson_and_matches_branch_total():
    import pipeline

    tour = pd.DataFrame(
        [
            {
                "來源單據號": "T001",
                "統一日期": "2026-05-01",
                "銷售點": "銅鑼灣分社",
                "銷售員": "Alice",
                "收款原幣金額": 100,
                "團負責人部門": "",
                "交易時間": "2026-05-01",
                "行程天數": 3,
                "數量": 1,
            },
            {
                "來源單據號": "T002",
                "統一日期": "2026-05-01",
                "銷售點": "銅鑼灣分社",
                "銷售員": "Ben",
                "收款原幣金額": 50,
                "團負責人部門": "郵輪部",
                "交易時間": "2026-05-01",
                "行程天數": 5,
                "數量": 1,
            },
        ]
    )
    others = pd.DataFrame(
        [
            {
                "來源單據號": "O001",
                "統一日期": "2026-05-01",
                "銷售點": "銅鑼灣分社",
                "銷售員": "Alice",
                "收款原幣金額": 25,
                "交易時間": "2026-05-01",
                "團名稱": "景點門票",
                "來源報表標籤": "門券all",
                "行程天數": 0,
                "數量": 1,
            },
            {
                "來源單據號": "O002",
                "統一日期": "2026-05-01",
                "銷售點": "銅鑼灣分社",
                "銷售員": "Ben",
                "收款原幣金額": 75,
                "交易時間": "2026-05-01",
                "團名稱": "景點門票",
                "來源報表標籤": "門券all",
                "行程天數": 0,
                "數量": 1,
            },
        ]
    )

    workbook_buffer, _, _ = pipeline.build_dashboard_data(
        tour,
        others,
        {"1A": "銅鑼灣分社", "2B": "專職銷售組"},
        ["銅鑼灣分社"],
        ["郵輪部"],
        ["Alice", "Ben"],
        include_branch_salesperson_sheet=True,
    )

    workbook = load_workbook(workbook_buffer, data_only=True)
    assert "分社經營統計_含銷售員" in workbook.sheetnames

    branch_rows = list(workbook["分社經營統計"].iter_rows(values_only=True))
    salesperson_rows = list(workbook["分社經營統計_含銷售員"].iter_rows(values_only=True))
    assert salesperson_rows[0] == (
        "文本",
        "單選",
        "銷售員",
        "日期",
        "月份",
        "旅行團",
        "郵輪",
        "票務",
        "旅行團交易人數",
        "票務交易數量",
    )

    branch_total = sum(float(row[4] or 0) + float(row[5] or 0) + float(row[6] or 0) for row in branch_rows[1:])
    salesperson_total = sum(
        float(row[5] or 0) + float(row[6] or 0) + float(row[7] or 0) for row in salesperson_rows[1:]
    )
    assert salesperson_total == branch_total == 250.0

    branch_tour_rows = list(workbook["分社每天旅行團交易人數"].iter_rows(values_only=True))
    branch_ticket_rows = list(workbook["分社每天票務交易數量"].iter_rows(values_only=True))
    branch_tour_people = sum(float(row[3] or 0) + float(row[4] or 0) for row in branch_tour_rows[1:])
    branch_ticket_qty = sum(float(row[4] or 0) for row in branch_ticket_rows[1:])
    salesperson_tour_people = sum(float(row[8] or 0) for row in salesperson_rows[1:])
    salesperson_ticket_qty = sum(float(row[9] or 0) for row in salesperson_rows[1:])

    assert salesperson_tour_people == branch_tour_people == 2.0
    assert salesperson_ticket_qty == branch_ticket_qty == 2.0
