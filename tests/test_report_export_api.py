from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.main import create_app
from backend.services import report_export_service


def test_dashboard_export_returns_xlsx_workbook(monkeypatch):
    monkeypatch.setattr(
        "backend.routers.exports.build_dashboard_report_workbook",
        lambda filters: b"dashboard-xlsx",
    )
    client = TestClient(create_app())

    response = client.post(
        "/api/exports/dashboard.xlsx",
        json={
            "years": [2026],
            "months": ["2026-06"],
            "dateRange": ["2026-06-01", "2026-06-30"],
            "branch": "全部分社",
            "salesGroup": "全部銷售組",
        },
    )

    assert response.status_code == 200
    assert response.headers["content-disposition"].startswith('attachment; filename="nbs_dashboard_report.xlsx"')
    assert response.content == b"dashboard-xlsx"


def test_quality_export_returns_xlsx_workbook(monkeypatch):
    monkeypatch.setattr("backend.routers.exports.build_quality_report_workbook", lambda: b"quality-xlsx")
    client = TestClient(create_app())

    response = client.get("/api/exports/quality.xlsx")

    assert response.status_code == 200
    assert response.headers["content-disposition"].startswith('attachment; filename="nbs_data_quality_scorecard.xlsx"')
    assert response.content == b"quality-xlsx"


def test_forecast_export_returns_xlsx_workbook(monkeypatch):
    monkeypatch.setattr("backend.routers.exports.build_forecast_report_workbook", lambda: b"forecast-xlsx")
    client = TestClient(create_app())

    response = client.get("/api/exports/forecast.xlsx")

    assert response.status_code == 200
    assert response.headers["content-disposition"].startswith('attachment; filename="nbs_forecast_report.xlsx"')
    assert response.content == b"forecast-xlsx"


def test_dashboard_report_workbook_contains_expected_sheets(monkeypatch):
    monkeypatch.setattr(
        "backend.services.report_export_service.build_dashboard_summary",
        lambda filters: {
            "kpis": [{"label": "淨營收", "value": "HKD 1", "delta": "+0", "note": "n", "accent": "blue"}],
            "revenueTotals": {
                "branchRevenue": 1.0,
                "specialistRevenue": 2.0,
                "combinedRevenue": 3.0,
                "formattedCombinedRevenue": "HKD 3",
                "scope": "Scope",
            },
            "dataFreshness": {
                "minDate": "2026-06-01",
                "maxDate": "2026-06-30",
                "rawRows": 3,
                "analysisRows": 2,
                "excludedRows": 1,
                "scope": "Scope",
            },
            "branchRanking": [{"rank": 1, "branch": "A", "tourRevenue": 1, "cruiseRevenue": 1, "ticketRevenue": 1, "totalRevenue": 3, "sharePct": 100}],
            "specialistRanking": [{"rank": 1, "specialist": "B", "tourRevenue": 1, "cruiseRevenue": 1, "ticketRevenue": 1, "totalRevenue": 3, "sharePct": 100}],
            "productMix": [{"product": "旅行團", "revenue": 3, "sharePct": 100}],
            "stabilityBaseline": {
                "name": "Baseline",
                "baselineMonth": "2026-05",
                "status": "matched",
                "formattedExpectedTotal": "HKD 3",
                "formattedActualTotal": "HKD 3",
                "expectedTotal": 3.0,
                "actualTotal": 3.0,
                "deltaAmount": 0.0,
                "deltaPct": 0.0,
                "summary": {"totalChecks": 1, "matchedChecks": 1, "driftChecks": 0},
                "checks": [],
                "coreValidation": {"status": "matched", "summary": {"totalChecks": 1, "matchedChecks": 1, "driftChecks": 0}, "checks": []},
                "freshnessUpdate": {"status": "stable", "summary": {"totalChecks": 1, "stableChecks": 1, "updatedChecks": 0}, "checks": []},
            },
            "scopeAudit": {},
            "revenueScope": "Scope",
            "exportReadiness": {},
        },
    )
    monkeypatch.setattr(
        "backend.services.report_export_service.build_dashboard_analytics",
        lambda filters: {
            "annualSummary": [{"year": 2026, "branchRevenue": 1, "specialistRevenue": 2, "combinedRevenue": 3, "branchSharePct": 33.33, "specialistSharePct": 66.67}],
            "monthlyTrend": [{"month": "2026-06", "branchRevenue": 1, "specialistRevenue": 2, "combinedRevenue": 3}],
            "branchRanking": [{"rank": 1, "branch": "A", "tourRevenue": 1, "cruiseRevenue": 1, "ticketRevenue": 1, "totalRevenue": 3, "sharePct": 100}],
            "specialistRanking": [{"rank": 1, "specialist": "B", "tourRevenue": 1, "cruiseRevenue": 1, "ticketRevenue": 1, "totalRevenue": 3, "sharePct": 100}],
            "productDrilldown": {
                "branch": [{"product": "旅行團", "revenue": 3, "sharePct": 100}],
                "specialist": [{"product": "旅行團", "revenue": 3, "sharePct": 100}],
            },
            "reconciliation": {"status": "matched", "combinedRevenue": 3.0, "checks": []},
            "revenueScope": "Scope",
            "appliedFilters": filters,
        },
    )

    content = report_export_service.build_dashboard_report_workbook({})
    workbook = load_workbook(BytesIO(content))

    assert "Dashboard KPIs" in workbook.sheetnames
    assert "Branch Ranking" in workbook.sheetnames
    assert "Monthly Trend" in workbook.sheetnames
    assert "Stability Baseline" in workbook.sheetnames
