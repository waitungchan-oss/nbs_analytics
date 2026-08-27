from pathlib import Path

import pandas as pd


def _candidate_builder(scope_id, intermediate):
    from io import BytesIO

    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "營收"
    sheet.append(["來源單據號", "收款原幣金額"])
    frame = intermediate.classified_tour if scope_id == "all" else intermediate.classified_tour.head(1)
    for _, row in frame.iterrows():
        sheet.append([row.get("來源單據號", ""), float(row.get("收款原幣金額", 0))])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _reference_builder(tour, others):
    from io import BytesIO

    from openpyxl import Workbook

    result = {}
    for key in ("ex", "ex_no_writeoff", "ex_no_writeoff_refund_transfer"):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "營收"
        sheet.append(["來源單據號", "收款原幣金額"])
        for _, row in tour.iterrows():
            sheet.append([row.get("來源單據號", ""), float(row.get("收款原幣金額", 0))])
        output = BytesIO()
        workbook.save(output)
        result[key] = output.getvalue()
    return result


def _frames():
    return (
        pd.DataFrame([{"來源單據號": "T-001", "統一日期": "2026-05-01", "收款原幣金額": 100}]),
        pd.DataFrame([{"來源單據號": "O-001", "統一日期": "2026-05-01", "收款原幣金額": 25}]),
    )


def test_fast_export_job_publishes_ready_manifest_after_equivalence(tmp_path):
    from backend.services.export_fast_path_service import build_fast_export_job

    tour, others = _frames()
    result = build_fast_export_job(
        tour,
        others,
        generation_token="generation-1",
        rules_fingerprint="rules-1",
        export_schema_version="schema-1",
        cache_root=tmp_path,
        reference_builder=_reference_builder,
        candidate_builder=_candidate_builder,
        worker_count=1,
    )

    assert result.status == "READY"
    assert result.manifest_path is not None
    assert result.fallback_reason is None
    assert result.timings["total_ms"] >= 0


def test_fast_export_job_uses_bounded_parallel_workers(tmp_path):
    from backend.services.export_fast_path_service import build_fast_export_job

    tour, others = _frames()
    result = build_fast_export_job(
        tour,
        others,
        generation_token="generation-parallel",
        rules_fingerprint="rules-1",
        export_schema_version="schema-1",
        cache_root=tmp_path,
        reference_builder=_reference_builder,
        candidate_builder=_candidate_builder,
        worker_count=3,
    )

    assert result.status == "READY"


def test_fast_export_job_falls_back_when_candidate_is_not_equivalent(tmp_path):
    from backend.services.export_fast_path_service import build_fast_export_job

    def mismatching_builder(scope_id, intermediate):
        return b"not-an-xlsx"

    tour, others = _frames()
    result = build_fast_export_job(
        tour,
        others,
        generation_token="generation-1",
        rules_fingerprint="rules-1",
        export_schema_version="schema-1",
        cache_root=tmp_path,
        reference_builder=_reference_builder,
        candidate_builder=mismatching_builder,
        worker_count=1,
    )

    assert result.status == "FALLBACK"
    assert result.manifest_path is None
    assert result.fallback_reason


def test_fast_export_job_blocks_publication_when_baseline_is_not_pass(tmp_path):
    from backend.services.export_fast_path_service import build_fast_export_job

    tour, others = _frames()
    result = build_fast_export_job(
        tour,
        others,
        generation_token="generation-baseline-blocked",
        rules_fingerprint="rules-1",
        export_schema_version="schema-1",
        cache_root=tmp_path,
        reference_builder=_reference_builder,
        candidate_builder=_candidate_builder,
        worker_count=1,
        baseline_status="DRIFT",
    )

    assert result.status == "FALLBACK"
    assert "baseline status" in (result.fallback_reason or "")


def test_rollout_mode_controls_whether_ready_fast_result_is_selected():
    from backend.services.export_fast_path_service import ExportRolloutMode, select_export_path

    assert select_export_path(ExportRolloutMode.DISABLED, fast_ready=True) == "legacy"
    assert select_export_path(ExportRolloutMode.SHADOW, fast_ready=True) == "legacy"
    assert select_export_path(ExportRolloutMode.OPT_IN, fast_ready=True) == "fast"
    assert select_export_path(ExportRolloutMode.DEFAULT, fast_ready=True) == "fast"
    assert select_export_path(ExportRolloutMode.DEFAULT, fast_ready=False) == "legacy"


def test_same_identity_reuses_trusted_reference_without_legacy_builder(tmp_path):
    from io import BytesIO

    from openpyxl import Workbook

    from backend.services.export_intermediate_service import ExportScope, build_scope_report_facts
    from backend.services.export_fast_path_service import build_fast_export_job_from_facts

    workbook = Workbook()
    workbook.active.title = "營收"
    workbook.active.append(["來源單據號", "收款原幣金額"])
    workbook.active.append(["T-001", 100])
    output = BytesIO()
    workbook.save(output)
    artifact = output.getvalue()
    tour, others = _frames()
    calls = []

    def facts_builder(intermediate):
        return {
            scope.value: build_scope_report_facts(intermediate, scope)
            for scope in ExportScope
        }

    def writer(_facts, path):
        path.write_bytes(artifact)

    def reference_builder(_tour, _others):
        calls.append("legacy")
        return {key: artifact for key in ("ex", "ex_no_writeoff", "ex_no_writeoff_refund_transfer")}

    kwargs = {
        "generation_token": "generation-cache-hit",
        "rules_fingerprint": "rules-1",
        "export_schema_version": "schema-1",
        "cache_root": tmp_path,
        "reference_builder": reference_builder,
        "facts_builder": facts_builder,
        "writer": writer,
        "worker_count": 1,
    }
    first = build_fast_export_job_from_facts(tour, others, **kwargs)
    second = build_fast_export_job_from_facts(tour, others, **kwargs)

    assert first.status == "READY"
    assert second.status == "READY"
    assert calls == ["legacy"]
    assert second.timings["reference_lookup_ms"] >= 0


def test_equivalence_failure_preserves_previous_trusted_reference_pointer(tmp_path):
    from io import BytesIO

    from openpyxl import Workbook

    from backend.services.export_intermediate_service import ExportScope, build_scope_report_facts
    from backend.services.export_fast_path_service import build_fast_export_job_from_facts

    def workbook_bytes(amount):
        workbook = Workbook()
        workbook.active.title = "營收"
        workbook.active.append(["來源單據號", "收款原幣金額"])
        workbook.active.append(["T-001", amount])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    reference_artifact = workbook_bytes(100)
    changed_artifact = workbook_bytes(101)
    candidate_artifact = reference_artifact
    tour, others = _frames()

    def facts_builder(intermediate):
        return {
            scope.value: build_scope_report_facts(intermediate, scope)
            for scope in ExportScope
        }

    def writer(_facts, path):
        path.write_bytes(candidate_artifact)

    def reference_builder(_tour, _others):
        return {key: reference_artifact for key in ("ex", "ex_no_writeoff", "ex_no_writeoff_refund_transfer")}

    kwargs = {
        "generation_token": "generation-pointer-safety",
        "rules_fingerprint": "rules-1",
        "export_schema_version": "schema-1",
        "cache_root": tmp_path,
        "reference_builder": reference_builder,
        "facts_builder": facts_builder,
        "writer": writer,
        "worker_count": 1,
    }
    first = build_fast_export_job_from_facts(tour, others, **kwargs)
    pointer = tmp_path / "trusted_reference" / "active.json"
    pointer_before = pointer.read_text(encoding="utf-8")

    candidate_artifact = changed_artifact
    second = build_fast_export_job_from_facts(tour, others, **kwargs)

    assert first.status == "READY"
    assert second.status == "FALLBACK"
    assert pointer.read_text(encoding="utf-8") == pointer_before


def test_manifest_failure_preserves_previous_trusted_reference_pointer(tmp_path, monkeypatch):
    from backend.services.export_intermediate_service import ExportScope, build_scope_report_facts
    from backend.services.export_fast_path_service import build_fast_export_job_from_facts
    import backend.services.export_fast_path_service as fast_service

    tour, others = _frames()
    from io import BytesIO
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.active.title = "營收"
    workbook.active.append(["來源單據號", "收款原幣金額"])
    workbook.active.append(["T-001", 100])
    output = BytesIO()
    workbook.save(output)
    candidate_artifact = {"value": output.getvalue()}

    def facts_builder(intermediate):
        return {
            scope.value: build_scope_report_facts(intermediate, scope)
            for scope in ExportScope
        }

    def writer(_facts, path):
        path.write_bytes(candidate_artifact["value"])

    def reference_builder(_tour, _others):
        return {key: candidate_artifact["value"] for key in ("ex", "ex_no_writeoff", "ex_no_writeoff_refund_transfer")}

    kwargs = {
        "generation_token": "generation-manifest-pointer-safety-1",
        "rules_fingerprint": "rules-1",
        "export_schema_version": "schema-1",
        "cache_root": tmp_path,
        "reference_builder": reference_builder,
        "facts_builder": facts_builder,
        "writer": writer,
        "worker_count": 1,
    }
    first = build_fast_export_job_from_facts(tour, others, **kwargs)
    pointer = tmp_path / "trusted_reference" / "active.json"
    pointer_before = pointer.read_text(encoding="utf-8")

    def fail_manifest(*_args, **_kwargs):
        raise OSError("manifest write failure")

    monkeypatch.setattr(fast_service, "publish_export_manifest", fail_manifest)
    kwargs["generation_token"] = "generation-manifest-pointer-safety-2"
    second = build_fast_export_job_from_facts(tour, others, **kwargs)

    assert first.status == "READY"
    assert second.status == "FALLBACK"
    assert pointer.read_text(encoding="utf-8") == pointer_before
