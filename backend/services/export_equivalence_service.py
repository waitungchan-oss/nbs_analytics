"""Semantic comparison for legacy and fast-path XLSX artifacts."""

from __future__ import annotations

import hashlib
import json
import math
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Mapping

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class CanonicalWorkbook:
    sheets: tuple[tuple[str, tuple[str, ...], tuple[tuple[object, ...], ...]], ...]
    schema_fingerprint: str
    data_fingerprint: str


@dataclass(frozen=True, slots=True)
class WorkbookEquivalenceReport:
    status: str
    schema_fingerprint: str
    data_fingerprint: str
    row_counts: Mapping[str, int]
    metric_summary: Mapping[str, object]
    mismatch_count: int
    mismatch_examples: tuple[Mapping[str, object], ...]


@dataclass(frozen=True, slots=True)
class ExportEquivalenceReport:
    status: str
    mismatch_count: int
    workbook_reports: Mapping[str, WorkbookEquivalenceReport]
    mismatch_examples: tuple[Mapping[str, object], ...]


def _value(value, *, money: bool = False):
    if value is None:
        return None
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if money:
        try:
            return str(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
        except Exception:
            return str(value).strip()
    if isinstance(value, bool):
        return bool(value)
    if isinstance(value, (int, float, Decimal)):
        return str(Decimal(str(value)).normalize())
    return str(value).strip()


def canonicalize_workbook(
    data: bytes,
    *,
    money_columns: tuple[str, ...] = (),
    stable_key_columns: tuple[str, ...] = (),
) -> CanonicalWorkbook:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheets = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        headers = tuple(str(value).strip() if value is not None else "" for value in (rows[0] if rows else ()))
        money_indexes = {index for index, header in enumerate(headers) if header in money_columns}
        canonical_rows = tuple(
            tuple(_value(value, money=index in money_indexes) for index, value in enumerate(row))
            for row in rows[1:]
        )
        if stable_key_columns and headers:
            key_indexes = tuple(headers.index(key) for key in stable_key_columns if key in headers)
            if key_indexes:
                canonical_rows = tuple(sorted(canonical_rows, key=lambda row: tuple(row[index] for index in key_indexes)))
        sheets.append((str(sheet_name), headers, canonical_rows))
    workbook.close()
    schema_payload = [(name, headers) for name, headers, _ in sheets]
    data_payload = [(name, headers, rows) for name, headers, rows in sheets]
    return CanonicalWorkbook(
        sheets=tuple(sheets),
        schema_fingerprint=hashlib.sha256(json.dumps(schema_payload, ensure_ascii=False, sort_keys=True).encode()).hexdigest(),
        data_fingerprint=hashlib.sha256(json.dumps(data_payload, ensure_ascii=False, sort_keys=True).encode()).hexdigest(),
    )


def build_workbook_metric_digest(
    data: bytes,
    *,
    money_columns: tuple[str, ...] = (),
    stable_key_columns: tuple[str, ...] = (),
) -> Mapping[str, object]:
    """Build bounded semantic metadata without retaining workbook rows."""
    canonical = canonicalize_workbook(
        data,
        money_columns=money_columns,
        stable_key_columns=stable_key_columns,
    )
    sheets = []
    for name, headers, rows in canonical.sheets:
        encoded_rows = json.dumps(
            rows, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        ).encode("utf-8")
        sheets.append({
            "name": name,
            "headers": list(headers),
            "row_count": len(rows),
            "row_fingerprint": hashlib.sha256(encoded_rows).hexdigest(),
        })
    return {
        "schema_fingerprint": canonical.schema_fingerprint,
        "sheets": sheets,
    }


def compare_export_digests(
    reference: Mapping[str, Mapping[str, object]],
    candidate: Mapping[str, Mapping[str, object]],
) -> bool:
    """Return true only when bounded semantic digest payloads match."""
    if set(reference) != set(candidate):
        return False
    return all(
        json.dumps(reference[key], ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        == json.dumps(candidate[key], ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        for key in reference
    )


def compare_workbooks(
    reference: bytes,
    candidate: bytes,
    *,
    money_columns: tuple[str, ...] = (),
    stable_key_columns: tuple[str, ...] = (),
) -> WorkbookEquivalenceReport:
    left = canonicalize_workbook(reference, money_columns=money_columns, stable_key_columns=stable_key_columns)
    right = canonicalize_workbook(candidate, money_columns=money_columns, stable_key_columns=stable_key_columns)
    examples: list[dict[str, object]] = []
    mismatch_count = 0
    if [(name, headers) for name, headers, _ in left.sheets] != [(name, headers) for name, headers, _ in right.sheets]:
        mismatch_count += 1
        examples.append({"kind": "schema", "reference": [(name, headers) for name, headers, _ in left.sheets], "candidate": [(name, headers) for name, headers, _ in right.sheets]})
    left_by_name = {name: (headers, rows) for name, headers, rows in left.sheets}
    right_by_name = {name: (headers, rows) for name, headers, rows in right.sheets}
    for name in sorted(set(left_by_name) & set(right_by_name)):
        left_headers, left_rows = left_by_name[name]
        right_headers, right_rows = right_by_name[name]
        if left_headers != right_headers:
            mismatch_count += 1
            if len(examples) < 20:
                examples.append({"sheet": name, "kind": "columns", "reference": left_headers, "candidate": right_headers})
            continue
        max_rows = max(len(left_rows), len(right_rows))
        for row_index in range(max_rows):
            left_row = left_rows[row_index] if row_index < len(left_rows) else None
            right_row = right_rows[row_index] if row_index < len(right_rows) else None
            max_columns = max(len(left_headers), len(right_headers))
            for column_index in range(max_columns):
                left_value = left_row[column_index] if left_row is not None and column_index < len(left_row) else None
                right_value = right_row[column_index] if right_row is not None and column_index < len(right_row) else None
                if left_value != right_value:
                    mismatch_count += 1
                    if len(examples) < 20:
                        examples.append({"sheet": name, "row": row_index + 2, "column": left_headers[column_index] if column_index < len(left_headers) else "", "reference": left_value, "candidate": right_value})
    row_counts = {name: len(rows) for name, _, rows in left.sheets}
    return WorkbookEquivalenceReport(
        status="PASS" if mismatch_count == 0 else "FAIL",
        schema_fingerprint=left.schema_fingerprint,
        data_fingerprint=left.data_fingerprint,
        row_counts=row_counts,
        metric_summary={"candidateDataFingerprint": right.data_fingerprint},
        mismatch_count=mismatch_count,
        mismatch_examples=tuple(examples),
    )


def compare_export_sets(reference: Mapping[str, bytes], candidate: Mapping[str, bytes]) -> ExportEquivalenceReport:
    examples: list[Mapping[str, object]] = []
    reports = {}
    missing = sorted(set(reference) ^ set(candidate))
    if missing:
        examples.append({"kind": "artifact_keys", "missing_or_extra": missing})
    for key in sorted(set(reference) & set(candidate)):
        reports[key] = compare_workbooks(reference[key], candidate[key])
        examples.extend({"artifact": key, **example} for example in reports[key].mismatch_examples)
    mismatch_count = (1 if missing else 0) + sum(report.mismatch_count for report in reports.values())
    return ExportEquivalenceReport(
        status="PASS" if mismatch_count == 0 else "FAIL",
        mismatch_count=mismatch_count,
        workbook_reports=reports,
        mismatch_examples=tuple(examples[:20]),
    )
