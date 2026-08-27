"""Bounded, atomic XLSX serialization for GMV report facts."""

from __future__ import annotations

import os
import shutil
import tempfile
import math
from threading import Event, Lock
import time
from concurrent.futures import ThreadPoolExecutor, TimeoutError, as_completed
from dataclasses import dataclass, replace
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Sequence

from openpyxl import load_workbook

from .gmv_export_intermediate_service import GmvReportFactSet, GmvReportFacts


DIMENSION_BY_LABEL = {"總退款": "total", "已退款": "paid"}
VALID_SCOPES = {"all", "no_writeoff", "official"}
CANONICAL_WORKBOOK_SCOPES = {
    "ex.xlsx": "all",
    "ex_no_writeoff.xlsx": "no_writeoff",
    "ex_no_writeoff_refund_transfer.xlsx": "official",
    "audit.xlsx": "official",
}
CANONICAL_SERIALIZER_SCOPES = (
    ("total", "ex.xlsx", "all"),
    ("total", "ex_no_writeoff.xlsx", "no_writeoff"),
    ("total", "ex_no_writeoff_refund_transfer.xlsx", "official"),
    ("paid", "ex.xlsx", "all"),
    ("paid", "ex_no_writeoff.xlsx", "no_writeoff"),
    ("paid", "ex_no_writeoff_refund_transfer.xlsx", "official"),
)
SERIALIZER_MIN_TIMEOUT_SECONDS = 120
SERIALIZER_MAX_TIMEOUT_SECONDS = 300
SERIALIZER_ROWS_PER_EXTRA_TIMEOUT_SECOND = 5_000


@dataclass(frozen=True, slots=True)
class SerializerJob:
    artifact_id: str
    facts: GmvReportFacts
    artifact_path: Path
    publication_gate: "SerializerPublicationGate"


@dataclass(frozen=True, slots=True)
class SerializerPublicationGate:
    equivalence_status: str
    checksum_status: str
    schema_status: str
    baseline_status: str
    shadow_status: str
    staging_only: bool = False

    @property
    def ready(self) -> bool:
        if self.staging_only:
            return all(status == "PASS" for status in (
                self.checksum_status, self.schema_status,
            )) and self.equivalence_status == "PENDING" and self.shadow_status == "PENDING"
        return all(status == "PASS" for status in (
            self.equivalence_status, self.checksum_status,
            self.schema_status, self.baseline_status, self.shadow_status,
        ))


@dataclass(frozen=True, slots=True)
class SerializerResult:
    artifact_id: str
    status: str
    path: Path
    bytes_written: int
    duration_ms: int
    error: str | None


def build_gmv_serializer_jobs(
    *,
    total_facts: GmvReportFactSet,
    paid_facts: GmvReportFactSet,
    staging_dir: Path,
    publication_gate: SerializerPublicationGate,
) -> tuple[SerializerJob, ...]:
    """Create the six canonical workbook jobs from facts only."""
    if total_facts.dimension != "總退款" or paid_facts.dimension != "已退款":
        raise ValueError("serializer fact sets have invalid refund dimensions")
    if not publication_gate.ready:
        raise ValueError("serializer publication gate is not ready")
    jobs: list[SerializerJob] = []
    facts_by_dimension = {"total": total_facts, "paid": paid_facts}
    for dimension, filename, scope_id in CANONICAL_SERIALIZER_SCOPES:
        fact_set = facts_by_dimension[dimension]
        facts = fact_set.facts_by_scope.get(scope_id)
        if facts is None or facts.scope_id != scope_id:
            raise ValueError(f"missing facts for {dimension}.{scope_id}")
        artifact_id = f"{dimension}.workbook.{filename}"
        jobs.append(SerializerJob(
            artifact_id=artifact_id,
            facts=facts,
            artifact_path=Path(staging_dir) / f"{dimension}-{filename}",
            publication_gate=publication_gate,
        ))
    return tuple(jobs)


def bounded_serializer_timeout_seconds(jobs: Sequence[SerializerJob]) -> int:
    """Give large production batches more time without allowing unbounded waits."""
    total_rows = sum(
        int(sum(int(count) for count in job.facts.row_counts.values()))
        for job in jobs
    )
    if total_rows <= 100_000:
        return SERIALIZER_MIN_TIMEOUT_SECONDS
    extra_rows = total_rows - 100_000
    extra_seconds = math.ceil(extra_rows / SERIALIZER_ROWS_PER_EXTRA_TIMEOUT_SECOND)
    return min(SERIALIZER_MAX_TIMEOUT_SECONDS, SERIALIZER_MIN_TIMEOUT_SECONDS + extra_seconds)


def _failure(artifact_id: str, path: Path, status: str, started: float, error: Exception) -> SerializerResult:
    return SerializerResult(
        artifact_id=artifact_id,
        status=status,
        path=path,
        bytes_written=0,
        duration_ms=round((time.perf_counter() - started) * 1000),
        error=f"{type(error).__name__}: {str(error)[:180]}",
    )


def serialize_gmv_report_facts(
    facts: GmvReportFacts,
    *,
    artifact_path: Path,
    writer: str = "openpyxl",
    artifact_id: str | None = None,
    publish_event: Event | None = None,
    publish_lock: Lock | None = None,
    published_ids: set[str] | None = None,
    publication_gate: SerializerPublicationGate | None = None,
) -> SerializerResult:
    started = time.perf_counter()
    path = Path(artifact_path)
    resolved_id = artifact_id or path.stem
    if writer != "openpyxl" or path.suffix.lower() != ".xlsx":
        return _failure(
            resolved_id, path, "SERIALIZER_INVALID_ARTIFACT", started,
            ValueError("GMV serializer requires openpyxl and .xlsx output"),
        )
    if publication_gate is None or not publication_gate.ready:
        return _failure(
            resolved_id, path, "SERIALIZER_GATE_FAILED", started,
            ValueError("GMV artifact publication gate is not fully PASS"),
        )
    if artifact_id is not None:
        parts = str(artifact_id).split(".", 2)
        canonical_scope = None
        if len(parts) == 3 and parts[1] == "workbook":
            canonical_scope = CANONICAL_WORKBOOK_SCOPES.get(parts[2])
        elif len(parts) == 2 and parts[0] in {"total", "paid"} and parts[1] in VALID_SCOPES:
            canonical_scope = parts[1]
        if len(parts) < 2 or parts[0] not in {"total", "paid"} or canonical_scope is None:
            return _failure(
                resolved_id, path, "SERIALIZER_INVALID_ARTIFACT", started,
                ValueError("invalid GMV artifact identity"),
            )
        if DIMENSION_BY_LABEL.get(str(facts.dimension)) != parts[0] or str(facts.scope_id) != canonical_scope:
            return _failure(
                resolved_id, path, "SERIALIZER_INVALID_ARTIFACT", started,
                ValueError("artifact identity does not match facts dimension or scope"),
            )
    try:
        if not facts.schema_fingerprint or not facts.data_fingerprint:
            raise ValueError("GMV facts require schema and data fingerprints")
        for sheet_name, frame in facts.sheets.items():
            columns = [str(column) for column in frame.columns]
            if not columns or len(columns) != len(set(columns)):
                raise ValueError(f"invalid schema for sheet {sheet_name}")
    except Exception as exc:
        return _failure(resolved_id, path, "SERIALIZER_INVALID_FACTS", started, exc)
    temporary_path: Path | None = None
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with NamedTemporaryFile(dir=path.parent, prefix=f".{path.name}.", suffix=".xlsx", delete=False) as handle:
            temporary_path = Path(handle.name)
        import pandas as pd

        with pd.ExcelWriter(temporary_path, engine=writer) as excel_writer:
            for sheet_name, frame in facts.sheets.items():
                frame.to_excel(excel_writer, sheet_name=str(sheet_name)[:31], index=False)
        with temporary_path.open("rb") as handle:
            os.fsync(handle.fileno())
        workbook = load_workbook(temporary_path, read_only=True, data_only=True)
        if list(workbook.sheetnames) != list(facts.sheets):
            workbook.close()
            raise ValueError("serialized workbook sheet contract mismatch")
        workbook.close()
        lock = publish_lock or Lock()
        with lock:
            if publish_event is not None and publish_event.is_set():
                raise TimeoutError("serializer publication cancelled")
            temporary_path.replace(path)
            if published_ids is not None:
                published_ids.add(resolved_id)
        size = path.stat().st_size
        return SerializerResult(resolved_id, "READY", path, size, round((time.perf_counter() - started) * 1000), None)
    except Exception as exc:
        return _failure(resolved_id, path, "SERIALIZER_EXCEPTION", started, exc)
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def serialize_gmv_workbooks_parallel(
    jobs: Sequence[SerializerJob],
    *,
    max_workers: int = 3,
    timeout_seconds: float | None = None,
) -> tuple[SerializerResult, ...]:
    if max_workers < 1:
        raise ValueError("max_workers must be >= 1")
    ordered_jobs = tuple(jobs)
    if not ordered_jobs:
        return ()
    artifact_ids = [job.artifact_id for job in ordered_jobs]
    if len(artifact_ids) != len(set(artifact_ids)):
        raise ValueError("artifact_id values must be unique")
    parents = {job.artifact_path.parent.resolve() for job in ordered_jobs}
    if len(parents) != 1:
        raise ValueError("parallel GMV artifacts must share one target directory")
    target_paths = [job.artifact_path.resolve() for job in ordered_jobs]
    if len(target_paths) != len(set(target_paths)):
        raise ValueError("parallel GMV artifact paths must be unique")
    staging_dir = Path(tempfile.mkdtemp(prefix=".gmv-serializer-", dir=next(iter(parents))))
    staged_jobs = tuple(replace(job, artifact_path=staging_dir / job.artifact_path.name) for job in ordered_jobs)
    results: dict[str, SerializerResult] = {}
    pool = ThreadPoolExecutor(
        max_workers=min(max_workers, len(ordered_jobs)), thread_name_prefix="gmv-serializer",
    )
    publish_events = {job.artifact_id: Event() for job in staged_jobs}
    publish_lock = Lock()
    published_ids: set[str] = set()
    futures = {
        pool.submit(
            serialize_gmv_report_facts,
            job.facts,
            artifact_path=job.artifact_path,
            artifact_id=job.artifact_id,
            publish_event=publish_events[job.artifact_id],
            publish_lock=publish_lock,
            published_ids=published_ids,
            publication_gate=job.publication_gate,
        ): job
        for job in staged_jobs
    }
    try:
        try:
            for future in as_completed(futures, timeout=timeout_seconds):
                job = futures[future]
                results[job.artifact_id] = future.result()
        except TimeoutError:
            with publish_lock:
                for future, job in futures.items():
                    if job.artifact_id not in results:
                        if job.artifact_id in published_ids:
                            results[job.artifact_id] = future.result()
                        else:
                            publish_events[job.artifact_id].set()
                            future.cancel()
                            results[job.artifact_id] = SerializerResult(
                                job.artifact_id, "SERIALIZER_TIMEOUT",
                                ordered_jobs[artifact_ids.index(job.artifact_id)].artifact_path, 0, 0,
                                "serializer timeout",
                            )
    finally:
        pool.shutdown(wait=True, cancel_futures=True)
    try:
        ordered_results = tuple(results[job.artifact_id] for job in staged_jobs)
        if not all(result.status == "READY" for result in ordered_results):
            return tuple(
                replace(result, path=original.artifact_path)
                for result, original in zip(ordered_results, ordered_jobs)
            )
        promoted: list[Path] = []
        backups: list[tuple[Path, Path]] = []
        try:
            for staged, original in zip(staged_jobs, ordered_jobs):
                if original.artifact_path.exists():
                    backup = staging_dir / f".backup-{original.artifact_path.name}"
                    original.artifact_path.replace(backup)
                    backups.append((original.artifact_path, backup))
                staged.artifact_path.replace(original.artifact_path)
                promoted.append(original.artifact_path)
        except Exception as exc:
            for path in promoted:
                path.unlink(missing_ok=True)
            for original, backup in reversed(backups):
                if backup.exists():
                    backup.replace(original)
            failure = _failure(
                "__batch__", ordered_jobs[0].artifact_path, "SERIALIZER_PUBLICATION_FAILED",
                time.perf_counter(), exc,
            )
            return tuple(
                replace(failure, artifact_id=job.artifact_id, path=job.artifact_path)
                for job in ordered_jobs
            )
        return tuple(
            replace(result, path=original.artifact_path)
            for result, original in zip(ordered_results, ordered_jobs)
        )
    finally:
        shutil.rmtree(staging_dir, ignore_errors=True)
