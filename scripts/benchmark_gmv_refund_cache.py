"""Read-only benchmark for GMV formal refund cache builds."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
import tempfile
import time
from contextlib import contextmanager
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Iterator

try:
    import resource
except ImportError:  # pragma: no cover - exercised on Windows runners
    resource = None  # type: ignore[assignment]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import database
import app_workflows
from backend.services.gmv_refund_repository import GmvRefundRepository
from backend.services.gmv_refund_service import RevenueFrames, build_gmv_formal_artifacts
from backend.services.gmv_export_cache_service import read_gmv_export_artifact


EXPECTED_ARTIFACT_COUNT = 11
LEGACY_WORKERS = 2


def _normalized_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return format(value, ".12g")
    return str(value).strip()


def _semantic_fingerprint(content: bytes, kind: str) -> str:
    digest = hashlib.sha256()

    def add(value: object) -> None:
        digest.update(json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8"))
        digest.update(b"\n")

    if kind == "xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            add({"sheet": sheet.title})
            for row in sheet.iter_rows(values_only=True):
                add([_normalized_cell(value) for value in row])
        workbook.close()
    elif kind == "csv":
        for row in csv.reader(StringIO(content.decode("utf-8-sig"))):
            add([_normalized_cell(value) for value in row])
    elif kind == "json":
        add(json.loads(content.decode("utf-8")))
    else:
        raise ValueError(f"unsupported artifact kind: {kind}")
    return digest.hexdigest()


def _validate_benchmark_cache_dir(cache_dir: str | Path) -> Path:
    root = Path(cache_dir).expanduser().resolve()
    lowered_parts = {part.lower() for part in root.parts}
    if ".nbs_runtime_cache" in lowered_parts or "nbs_runtime_cache" in lowered_parts:
        raise ValueError("benchmark cache_dir cannot be a formal runtime cache")
    temp_root = Path(tempfile.gettempdir()).resolve()
    approved_roots = (temp_root, PROJECT_ROOT / ".nbs_agent_runtime" / "benchmarks")
    if not any(root == allowed or allowed in root.parents for allowed in approved_roots):
        raise ValueError("benchmark cache_dir must be under a temporary or isolated benchmark root")
    if root == Path.cwd().resolve() or root == temp_root or root == PROJECT_ROOT / ".nbs_runtime_cache":
        raise ValueError("benchmark cache_dir is too broad")
    return root


@contextmanager
def _timed_patch(module: Any, name: str, timings: dict[str, list[float]]) -> Iterator[None]:
    original = getattr(module, name)

    def wrapped(*args: Any, **kwargs: Any) -> Any:
        started = time.perf_counter()
        try:
            return original(*args, **kwargs)
        finally:
            timings.setdefault(name, []).append((time.perf_counter() - started) * 1000)

    setattr(module, name, wrapped)
    try:
        yield
    finally:
        setattr(module, name, original)


def _peak_rss_bytes() -> int | None:
    if resource is None:
        return None
    value = int(resource.getrusage(resource.RUSAGE_SELF).ru_maxrss)
    if sys.platform == "darwin":
        return value
    return value * 1024


def _prepare_frames(db_path: str | Path) -> tuple[RevenueFrames, float]:
    started = time.perf_counter()
    raw_tour, raw_others = database.load_all_data_from_db(db_path=db_path)
    formal_tour, formal_others, _ = app_workflows._build_revenue_scope_frames(raw_tour, raw_others)
    return (
        RevenueFrames(raw_tour, raw_others, formal_tour, formal_others),
        (time.perf_counter() - started) * 1000,
    )


def run_gmv_cache_benchmark(
    *,
    db_path: str | Path,
    version_id: str,
    cache_dir: str | Path,
    mode: str = "legacy",
    workers: int = LEGACY_WORKERS,
) -> dict[str, object]:
    """Benchmark one cache build without mutating SQLite or formal cache."""
    if mode != "legacy":
        raise ValueError(f"unsupported benchmark mode: {mode}")
    if workers != LEGACY_WORKERS:
        raise ValueError(f"legacy benchmark requires workers={LEGACY_WORKERS}")

    cache_root = _validate_benchmark_cache_dir(cache_dir)
    repository = GmvRefundRepository(db_path)
    active = repository.load_active_scope()
    if active is None or str(active["version_id"]) != str(version_id):
        raise ValueError("version_id is not the active GMV version")
    frames, base_ms = _prepare_frames(db_path)
    timings: dict[str, list[float]] = {}
    started = time.perf_counter()
    import backend.services.gmv_export_cache_service as cache_service

    patches = [
        _timed_patch(app_workflows, "_apply_gmv_refund_adjustments", timings),
        _timed_patch(app_workflows, "_gmv_summary_rows", timings),
        _timed_patch(app_workflows, "build_formal_gmv_workbooks", timings),
        _timed_patch(app_workflows, "_compute_gmv_exclusion_workbooks", timings),
        _timed_patch(cache_service, "build_gmv_export_cache", timings),
    ]
    try:
        for patch in patches:
            patch.__enter__()
        result = build_gmv_formal_artifacts(
            repository=repository,
            version_id=str(version_id),
            revenue_frames=frames,
            rule_version=str(active["rule_version"]),
            cache_dir=cache_root,
        )
    finally:
        for patch in reversed(patches):
            patch.__exit__(None, None, None)

    total_ms = (time.perf_counter() - started) * 1000 + base_ms
    manifest = result.cache_manifest
    if len(manifest.artifacts) != EXPECTED_ARTIFACT_COUNT:
        raise AssertionError(f"legacy artifact contract changed: {len(manifest.artifacts)}")
    equivalence_started = time.perf_counter()
    fingerprints = {
        key: _semantic_fingerprint(
            read_gmv_export_artifact(manifest, cache_root, key),
            str(record["kind"]),
        )
        for key, record in manifest.artifacts.items()
    }
    equivalence_ms = (time.perf_counter() - equivalence_started) * 1000
    artifact_bytes = sum(int(record.get("bytes", 0)) for record in manifest.artifacts.values())
    return {
        "mode": mode,
        "workers": workers,
        "basePreparationMs": round(base_ms, 1),
        "adjustmentMs": round(sum(timings.get("_apply_gmv_refund_adjustments", [])), 1),
        "factsMs": round(sum(timings.get("_gmv_summary_rows", [])), 1),
        "serializationMs": round(
            sum(timings.get("build_formal_gmv_workbooks", []))
            + max(timings.get("_compute_gmv_exclusion_workbooks", [0.0])),
            1,
        ),
        "equivalenceMs": round(equivalence_ms, 1),
        "fingerprintMs": round(equivalence_ms, 1),
        "cacheWriteMs": round(sum(timings.get("build_gmv_export_cache", [])), 1),
        "totalMs": round(total_ms, 1),
        "artifactBytes": artifact_bytes,
        "artifactCount": len(manifest.artifacts),
        "equivalenceStatus": "BASELINE_FINGERPRINT_CAPTURED",
        "artifactFingerprints": fingerprints,
        "peakRssBytes": _peak_rss_bytes(),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--db-path", required=True)
    parser.add_argument("--version-id", required=True)
    parser.add_argument("--cache-dir", required=True)
    parser.add_argument("--mode", default="legacy", choices=("legacy",))
    parser.add_argument("--workers", type=int, default=LEGACY_WORKERS)
    parser.add_argument("--output")
    args = parser.parse_args()
    payload = run_gmv_cache_benchmark(
        db_path=args.db_path,
        version_id=args.version_id,
        cache_dir=args.cache_dir,
        mode=args.mode,
        workers=args.workers,
    )
    rendered = json.dumps(payload, ensure_ascii=False, indent=2)
    if args.output:
        cache_root = _validate_benchmark_cache_dir(args.cache_dir)
        output_path = Path(args.output).expanduser().resolve()
        output_path.relative_to(cache_root)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(rendered + "\n", encoding="utf-8")
    print(rendered)


if __name__ == "__main__":
    main()
